import openpyxl
import sqlite3
import os

def import_rates():
    # Διαδρομή αρχείου
    excel_file = os.path.join("data", "rates.xlsx")
    db_file = "taxi.db"

    if not os.path.exists(excel_file):
        print(f"❌ Το αρχείο '{excel_file}' δεν βρέθηκε!")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        print(f"📂 Επεξεργασία αρχείου: {excel_file}...")

        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()

        # ---------------------------------------------------------
        # ΒΗΜΑ 1: Έξυπνη Εισαγωγή Προορισμών
        # ---------------------------------------------------------
        dest_map = {} 
        for col_idx, cell in enumerate(sheet.iter_cols(min_col=2, max_row=1, values_only=True), start=2):
            dest_name = cell[0]
            if dest_name:
                dest_name = str(dest_name).strip()
                cursor.execute("SELECT id FROM destinations WHERE name = ?", (dest_name,))
                existing = cursor.fetchone()

                if existing:
                    dest_id = existing[0]
                else:
                    cursor.execute("INSERT INTO destinations (name, is_active, sort_order) VALUES (?, 1, 99)", (dest_name,))
                    dest_id = cursor.lastrowid
                    print(f"   ➕ ΝΕΟΣ Προορισμός: {dest_name}")
                
                dest_map[col_idx] = dest_id

        # ---------------------------------------------------------
        # ΒΗΜΑ 2: Έξυπνη Εισαγωγή Ξενοδοχείων
        # ---------------------------------------------------------
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            hotel_name = row[0]
            if hotel_name:
                hotel_name = str(hotel_name).strip()
                cursor.execute("SELECT id FROM hotels WHERE name = ?", (hotel_name,))
                existing_hotel = cursor.fetchone()

                if existing_hotel:
                    hotel_id = existing_hotel[0]
                else:
                    cursor.execute("INSERT INTO hotels (name, is_active, sort_order) VALUES (?, 1, 99)", (hotel_name,))
                    hotel_id = cursor.lastrowid
                    print(f"   ➕ ΝΕΟ Ξενοδοχείο: {hotel_name}")

                # -----------------------------------------------------
                # ΒΗΜΑ 3: ΑΣΦΑΛΗΣ Ενημέρωση Credits (Check -> Update/Insert)
                # -----------------------------------------------------
                for i in range(1, len(row)):
                    col_idx = i + 1
                    if col_idx in dest_map:
                        dest_id = dest_map[col_idx]
                        val = row[i]
                        
                        try:
                            credits_val = int(val) if val is not None else 0
                        except ValueError:
                            credits_val = 0

                        # 🔥 Η ΑΛΛΑΓΗ ΕΙΝΑΙ ΕΔΩ: Ελέγχουμε αν υπάρχει τιμή
                        cursor.execute("SELECT credits FROM rates WHERE hotel_id=? AND dest_id=?", (hotel_id, dest_id))
                        existing_rate = cursor.fetchone()

                        if existing_rate:
                            # Αν υπάρχει, κάνουμε UPDATE
                            cursor.execute("""
                                UPDATE rates SET credits=? 
                                WHERE hotel_id=? AND dest_id=?
                            """, (credits_val, hotel_id, dest_id))
                        else:
                            # Αν δεν υπάρχει, κάνουμε INSERT
                            cursor.execute("""
                                INSERT INTO rates (hotel_id, dest_id, credits) 
                                VALUES (?, ?, ?)
                            """, (hotel_id, dest_id, credits_val))

        conn.commit()
        conn.close()
        print("\n✅ Ο συγχρονισμός ολοκληρώθηκε επιτυχώς!")
        print("   - Οι τιμές ενημερώθηκαν χωρίς να δημιουργηθούν διπλότυπα.")

    except Exception as e:
        print(f"❌ Σφάλμα: {e}")

if __name__ == "__main__":
    import_rates()
