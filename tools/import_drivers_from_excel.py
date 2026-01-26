import openpyxl
import sqlite3
import os

def import_drivers_update():
    # Καθορισμός διαδρομών
    tools_path = os.path.dirname(os.path.abspath(__file__))
    base_path = os.path.abspath(os.path.join(tools_path, ".."))
    excel_file = os.path.join(base_path, "data", "drivers.xlsx")
    db_path = os.path.join(base_path, "taxi.db")
    print(f"🔍 Έλεγχος στο: {excel_file}") # Προσθήκη για να βλέπουμε πού ψάχνει

    # Έλεγχος αρχείου
    if not os.path.exists(excel_file):
        # Fallback: Ψάξε στον τρέχοντα φάκελο αν δεν είναι στο /data
        excel_file = os.path.join(base_path, "drivers.xlsx")
        if not os.path.exists(excel_file):
            print(f"❌ Το αρχείο drivers.xlsx δεν βρέθηκε ούτε στο /data ούτε στον φάκελο της εφαρμογής!")
            return

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        print(f"🚀 Έναρξη ενημέρωσης από: {excel_file}")
        
        updated_count = 0
        inserted_count = 0

        # Διαβάζουμε από τη γραμμή 2
        for row_idx in range(2, ws.max_row + 1):
            col_new = str(ws.cell(row_idx, 1).value or "").strip()
            col_old = str(ws.cell(row_idx, 2).value or "").strip()
            name = str(ws.cell(row_idx, 3).value or "").strip().upper()
            phone = str(ws.cell(row_idx, 4).value or "").strip()

            if not name and not col_new:
                continue  # Προσπέρασε κενές γραμμές

            found_id = None
            
            # Στρατηγική Αναζήτησης:
            # 1. Ψάχνουμε με τον μοναδικό κωδικό 'col_new' (αν υπάρχει)
            if col_new:
                cursor.execute("SELECT id FROM drivers WHERE col_new = ?", (col_new,))
                res = cursor.fetchone()
                if res:
                    found_id = res[0]
            
            # 2. Αν δεν βρέθηκε με κωδικό, ψάχνουμε με το 'Όνομα'
            # (Χρήσιμο για πρώτη φορά ή αν δεν έχουν κωδικό)
            if not found_id and name:
                cursor.execute("SELECT id FROM drivers WHERE UPPER(name) = ?", (name,))
                res = cursor.fetchone()
                if res:
                    found_id = res[0]

            if found_id:
                # UPDATE: Ενημερώνουμε τα πάντα για να είμαστε σίγουροι
                cursor.execute("""
                    UPDATE drivers 
                    SET col_new=?, col_old=?, name=?, phone=? 
                    WHERE id=?
                """, (col_new, col_old, name, phone, found_id))
                updated_count += 1
                # print(f"   ✏️ Ενημέρωση: {name}")
            else:
                # INSERT: Νέος οδηγός
                cursor.execute("""
                    INSERT INTO drivers (col_new, col_old, name, phone) 
                    VALUES (?, ?, ?, ?)
                """, (col_new, col_old, name, phone))
                inserted_count += 1
                print(f"   ➕ ΝΕΟΣ Οδηγός: {name}")

        conn.commit()
        conn.close()
        
        print("\n✅ Ολοκληρώθηκε!")
        print(f"   - Ενημερώθηκαν (Updated): {updated_count}")
        print(f"   - Προστέθηκαν (Inserted): {inserted_count}")

    except Exception as e:
        print(f"❌ Σφάλμα: {e}")

if __name__ == "__main__":
    import_drivers_update()
