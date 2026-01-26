# --- 8. Εξαγωγή Rates (Export) ---
from flask import Blueprint, render_template, request, redirect, url_for, Response
from datetime import datetime
import io 
import csv
from utils import get_db_connection, get_setting, log_action 

# Δημιουργία του Blueprint
exportrates_bp = Blueprint('exportrates', __name__)

@exportrates_bp.route('/export_rates', methods=['GET'])
def export_rates():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "Σφάλμα: Η βιβλιοθήκη openpyxl δεν είναι εγκατεστημένη. Τρέξε 'pip install openpyxl'", 500

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM destinations WHERE is_active=1 ORDER BY sort_order, name")
    destinations = cursor.fetchall()
    cursor.execute("SELECT id, name FROM hotels WHERE is_active=1 ORDER BY sort_order, name")
    hotels = cursor.fetchall()
    cursor.execute("SELECT hotel_id, dest_id, credits FROM rates")
    rates = cursor.fetchall()
    conn.close()
    
    rates_dict = {}
    for r in rates:
        rates_dict[(r['hotel_id'], r['dest_id'])] = int(r['credits'])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rates"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")
    
    ws['A1'] = "Hotel"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A1'].border = thin_border
    
    for col_idx, dest in enumerate(destinations, start=2):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = dest['name']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, hotel in enumerate(hotels, start=2):
        cell_h = ws.cell(row=row_idx, column=1)
        cell_h.value = hotel['name']
        cell_h.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell_h.font = Font(bold=True)
        cell_h.border = thin_border
        
        for col_idx, dest in enumerate(destinations, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = rates_dict.get((hotel['id'], dest['id']), 0)
            cell.value = val if val != 0 else ""
            cell.alignment = center_align
            cell.border = thin_border
            
    ws.column_dimensions['A'].width = 25
    for i in range(2, len(destinations) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"export_rates_{timestamp}.xlsx"
    
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"})